import io
import re

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from plex_data import (
    fetch_plex_sheets,
    get_credentials,
    get_or_create_worksheet,
    is_odbc_configured,
    is_registry_configured,
    load_project_registry,
    parse_part_numbers,
)

STATUS_COLORS = {"Profit": "#0ca30c", "Loss": "#d03b3b"}
STATUS_COLORS_HEX = {"Profit": "0CA30C", "Loss": "D03B3B"}  # no '#', for openpyxl

PO_REGISTRY_HEADERS = ["Okay PN", "Total PO $"]


def get_po_registry_worksheet():
    return get_or_create_worksheet("PO Registry", PO_REGISTRY_HEADERS, rows=1000)


def load_po_registry() -> dict[str, float]:
    """Returns {Okay PN: Total PO $} remembered from previous days. Returns an
    empty registry (rather than raising) if Google Sheets isn't configured, so
    the app still works without it — see PO_REGISTRY_SETUP.md. Fetches
    UNFORMATTED_VALUE so currency-formatted cells (e.g. '$174,827.00', or
    '$ -' for zero — both just display formatting on an underlying number)
    come back as plain numbers instead of strings that fail float(). Rows
    with a genuinely blank/non-numeric Total PO $ (e.g. a new Okay PN added
    by hand and not yet filled in) are skipped individually rather than
    blanking the whole registry."""
    try:
        worksheet = get_po_registry_worksheet()
        records = worksheet.get_all_records(value_render_option="UNFORMATTED_VALUE")
    except Exception:
        return {}

    registry = {}
    for r in records:
        pn = str(r.get("Okay PN", "")).strip()
        if not pn:
            continue
        try:
            registry[normalize_pn(pn)] = float(r["Total PO $"])
        except (KeyError, TypeError, ValueError):
            continue
    return registry


GDRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


@st.cache_resource
def get_drive_session():
    """An authenticated HTTP session for the Google Drive API, reusing the same
    service account as the PO/Project registries — just needs Drive scope added."""
    from google.auth.transport.requests import AuthorizedSession

    return AuthorizedSession(get_credentials(GDRIVE_SCOPES))


def is_drive_configured() -> bool:
    """True when a Drive folder is configured to read the daily Plex export
    from (daily_plex_export.py writes there) — the fallback source used when
    ODBC isn't available locally, e.g. on a Cloud deployment."""
    try:
        return "gcp_service_account" in st.secrets and "plex_export_folder_id" in st.secrets
    except Exception:
        return False


def fetch_drive_files() -> list[io.BytesIO]:
    """Downloads every CSV/Excel file in the configured Google Drive folder —
    populated daily by daily_plex_export.py — returning file-like objects with
    a `.name` attribute, a drop-in replacement for Streamlit's uploaded-file
    objects, compatible with load_all_sheets()."""
    session = get_drive_session()
    folder_id = st.secrets["plex_export_folder_id"]
    response = session.get(
        "https://www.googleapis.com/drive/v3/files",
        params={
            "q": f"'{folder_id}' in parents and trashed = false",
            "fields": "files(id, name)",
        },
    )
    response.raise_for_status()
    files = response.json().get("files", [])

    buffers = []
    for item in files:
        if not item["name"].lower().endswith((".csv", ".xlsx", ".xls")):
            continue
        response = session.get(
            f"https://www.googleapis.com/drive/v3/files/{item['id']}", params={"alt": "media"}
        )
        response.raise_for_status()
        buffer = io.BytesIO(response.content)
        buffer.name = item["name"]
        buffers.append(buffer)
    return buffers


def build_prefix_balances(final_sheets: dict[str, pd.DataFrame]) -> dict[str, float]:
    """Sums each row's Budget Left across every uploaded file, grouped by
    Okay PN prefix (same prefix used for the Step 4 'Project' label)."""
    combined = pd.concat(
        [df[df[df.columns[0]] != "TOTAL"] for df in final_sheets.values()],
        ignore_index=True,
    )
    prefixes = combined["Okay PN"].dropna().astype(str).map(extract_pn_prefix)
    return combined.assign(_prefix=prefixes).groupby("_prefix")["Budget Left"].sum().to_dict()


def build_open_projects(final_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Joins the hand-maintained Project Registry with a live-computed
    Project Balance (sum of Budget Left for that row's Part Number(s))."""
    registry_df = load_project_registry()
    if registry_df.empty:
        return registry_df.assign(**{"Project Balance": pd.Series(dtype=float)})

    prefix_balances = build_prefix_balances(final_sheets)
    registry_df["Project Balance"] = registry_df["Part Number"].apply(
        lambda pn: sum(prefix_balances.get(n, 0.0) for n in parse_part_numbers(pn))
    )
    ordered_cols = [
        "Customer",
        "Project",
        "Part Number",
        "Project Balance",
        "Contingency/Management Reserve Used",
        "Expected Project End Date",
        "NOTES",
    ]
    return registry_df[ordered_cols]


HISTORY_HEADERS = ["Date", "Customer", "Project", "Part Number", "Project Balance"]
HISTORY_KEY_COLS = ["Customer", "Project", "Part Number"]


def get_history_worksheet():
    return get_or_create_worksheet("History", HISTORY_HEADERS, rows=2000)


def load_history() -> pd.DataFrame:
    """Returns every logged Project Balance snapshot, one row per Date x
    Customer/Project/Part Number. Fetches UNFORMATTED_VALUE and coerces
    Project Balance to numeric (dropping rows that fail) so a stray
    currency-formatted or blank cell can't break the trend chart or, worse,
    silently empty the whole history via save_history's overwrite — same
    class of bug fixed in load_po_registry. Returns an empty frame if Google
    Sheets isn't configured."""
    try:
        worksheet = get_history_worksheet()
        records = worksheet.get_all_records(value_render_option="UNFORMATTED_VALUE")
    except Exception:
        return pd.DataFrame(columns=HISTORY_HEADERS)

    if not records:
        return pd.DataFrame(columns=HISTORY_HEADERS)
    history = pd.DataFrame(records)
    history["Project Balance"] = pd.to_numeric(history["Project Balance"], errors="coerce")
    return history.dropna(subset=["Project Balance"])


def log_open_projects_snapshot(open_projects: pd.DataFrame) -> bool:
    """Appends today's Project Balance for each Open Projects row to the
    History tab, so trends can be charted over time. Re-running Generate
    Final Results on the same day replaces that day's rows instead of
    duplicating them. Returns False (without raising) if Sheets isn't
    configured or there's nothing to log."""
    if open_projects.empty:
        return False
    try:
        today = pd.Timestamp.now().strftime("%Y-%m-%d")
        history = load_history()
        history = history[history["Date"] != today]

        new_rows = open_projects[HISTORY_KEY_COLS + ["Project Balance"]].copy()
        new_rows.insert(0, "Date", today)

        combined = pd.concat([history, new_rows], ignore_index=True)[HISTORY_HEADERS]
        worksheet = get_history_worksheet()
        worksheet.clear()
        worksheet.update([HISTORY_HEADERS] + combined.values.tolist())
        return True
    except Exception:
        return False


st.set_page_config(page_title="Tooling Data Cleaning & Budget Tool", layout="wide")

REQUIRED_COLUMNS = [
    "Tooling Line Item Description",
    "Okay PN",
    "Tooling Job No.",
    "Total Revenue",
    "Invoiced Revenue",
    "Vendor POs Cost",
    "Labor Cost",
    "Total Cost",
    "Profit or Loss",
]

SUM_COLS = ["Invoiced Revenue", "Vendor POs Cost", "Labor Cost", "Total Cost"]

NUMERIC_TOTAL_COLS = [
    "Total Revenue",
    "Invoiced Revenue",
    "Vendor POs Cost",
    "Labor Cost",
    "Total Cost",
    "Profit or Loss",
    "Total PO $",
    "Budget Left",
]


def normalize_pn(value) -> str:
    """Normalizes Okay PN for registry matching: trims whitespace, collapses
    internal whitespace, strips a trailing dash (Plex exports sometimes add
    one, e.g. '932 A PD-01-' vs '932 A PD-01'), and ignores case."""
    text = re.sub(r"\s+", " ", str(value).strip().upper())
    return text.rstrip("- ")


def is_junk_okay_pn(value) -> bool:
    """True for an Okay PN that's just a bare number/revision code with no
    real description — e.g. '931-01', '924-1 -01', or a number plus a
    standalone 'ASY' suffix like '938 ASY-01'. These rows get dropped during
    cleaning. A PN with real text after ASY (e.g. '4243 ASY A MC-01-01') is
    kept — only the first, whole-word 'ASY' is stripped before checking."""
    text = re.sub(r"(?<![A-Z])ASY(?![A-Z])", "", str(value).strip().upper(), count=1)
    return bool(re.fullmatch(r"[\d\s\-]*", text))


EXCLUDED_OKAY_PNS = {"4066M-01", "388M-01", "388-1P-03", "388-1P-02", "1/1/4243"}


def consolidate_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    """Merge rows that share the same Okay PN, following the same rules as the
    original script: sum Total Revenue only when it differs across the group,
    zero out the other cost/revenue columns on the extra rows (or sum them
    onto the first row when they differ), then drop the extra rows."""
    df_cleaned = df[REQUIRED_COLUMNS].copy()
    df_work = df_cleaned.copy()
    rows_to_drop = []

    for okay_pn in df_work["Okay PN"].unique():
        group_indices = df_work[df_work["Okay PN"] == okay_pn].index.tolist()
        if len(group_indices) <= 1:
            continue

        group = df_work.loc[group_indices]
        first_idx, other_indices = group_indices[0], group_indices[1:]

        if len(group["Total Revenue"].unique()) != 1:
            df_work.loc[first_idx, "Total Revenue"] = group["Total Revenue"].sum()

        for col in SUM_COLS:
            if len(group[col].unique()) == 1:
                df_work.loc[other_indices, col] = 0
            else:
                df_work.loc[first_idx, col] = group[col].sum()

        rows_to_drop.extend(other_indices)

    df_cleaned = df_work.drop(rows_to_drop).reset_index(drop=True)
    df_cleaned["Profit or Loss"] = df_cleaned["Total Revenue"] - df_cleaned["Total Cost"]
    return df_cleaned


def load_all_sheets(uploaded_files) -> dict[str, pd.DataFrame]:
    sheets = {}
    for uploaded_file in uploaded_files:
        name = uploaded_file.name
        if name.lower().endswith(".csv"):
            sheets[name.rsplit(".", 1)[0]] = pd.read_csv(uploaded_file)
        elif name.lower().endswith((".xlsx", ".xls")):
            excel_file = pd.ExcelFile(uploaded_file)
            for sheet in excel_file.sheet_names:
                sheets[f"{name}_{sheet}"] = pd.read_excel(excel_file, sheet_name=sheet)
    return sheets


def add_totals_row(df: pd.DataFrame) -> pd.DataFrame:
    totals = {col: (df[col].sum() if col in NUMERIC_TOTAL_COLS else "") for col in df.columns}
    totals[df.columns[0]] = "TOTAL"
    return pd.concat([df, pd.DataFrame([totals])], ignore_index=True)


def extract_pn_prefix(value: str) -> str:
    """Leading run of letters/digits, stopping at the first dash, dot, space,
    underscore, etc. — e.g. '924-1' and '924-01' both become '924'."""
    match = re.match(r"^[A-Za-z0-9]+", value)
    return match.group(0) if match else value


def derive_sheet_label(df: pd.DataFrame, fallback: str) -> str:
    """Human-readable heading for a sheet, e.g. '932/4066', based on its Okay
    PN prefixes — used instead of the raw uploaded filename, which is often an
    auto-generated export name like 'Query_2026_07_10-09-12-16'. Falls back to
    that filename if no prefixes can be derived (e.g. an empty sheet)."""
    data_rows = df[df[df.columns[0]] != "TOTAL"]
    prefixes = data_rows["Okay PN"].dropna().astype(str).map(extract_pn_prefix).unique()
    return "/".join(prefixes) if len(prefixes) else fallback


def build_project_summary(final_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """One row per uploaded file, reusing each file's existing TOTAL row from
    Step 3. If a file contains multiple unique Tooling Job No. (or Okay PN
    prefix) values, they are joined with '/'."""
    rows = []
    for sheet_name, df in final_sheets.items():
        data_rows = df[df[df.columns[0]] != "TOTAL"]
        totals_row = df.iloc[-1]
        job_numbers = data_rows["Tooling Job No."].dropna().astype(str).unique()
        pn_prefixes = data_rows["Okay PN"].dropna().astype(str).map(extract_pn_prefix).unique()

        rows.append(
            {
                "File": sheet_name,
                "Project": "/".join(pn_prefixes),
                "Tooling Job No.": "/".join(job_numbers),
                "Total PO $": totals_row["Total PO $"],
                "Total Cost": totals_row["Total Cost"],
            }
        )

    summary = pd.DataFrame(rows)
    summary["Profit or Loss ($)"] = summary["Total PO $"] - summary["Total Cost"]
    summary["Status"] = summary["Profit or Loss ($)"].apply(lambda v: "Profit" if v >= 0 else "Loss")
    return summary


def build_excel(final_sheets: dict[str, pd.DataFrame]) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df_final in final_sheets.items():
            df_final.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    return buffer.getvalue()


def build_project_summary_excel(project_summary: pd.DataFrame) -> bytes:
    """Project summary table plus a native, editable Excel bar chart colored
    green/red by Profit/Loss status."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        project_summary.to_excel(writer, sheet_name="Project Summary", index=False)
        worksheet = writer.sheets["Project Summary"]
        n_rows = len(project_summary)

        header_fill = PatternFill("solid", fgColor="2A78D6")
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        currency_cols = ["Total PO $", "Total Cost", "Profit or Loss ($)"]
        for col_name in currency_cols:
            col_letter = get_column_letter(project_summary.columns.get_loc(col_name) + 1)
            for row in range(2, n_rows + 2):
                worksheet[f"{col_letter}{row}"].number_format = '"$"#,##0.00'

        for i, col_name in enumerate(project_summary.columns, start=1):
            max_len = max(project_summary[col_name].astype(str).map(len).max(), len(col_name)) + 2
            worksheet.column_dimensions[get_column_letter(i)].width = max_len

        chart = BarChart()
        chart.type = "col"
        chart.title = "Profit or Loss by Project"
        chart.y_axis.title = "Profit or Loss ($)"
        chart.x_axis.title = "Project"
        chart.legend = None
        chart.height, chart.width = 10, 20

        profit_col = project_summary.columns.get_loc("Profit or Loss ($)") + 1
        project_col = project_summary.columns.get_loc("Project") + 1
        data = Reference(worksheet, min_col=profit_col, min_row=1, max_row=n_rows + 1)
        cats = Reference(worksheet, min_col=project_col, min_row=2, max_row=n_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        series = chart.series[0]
        series.data_points = [DataPoint(idx=i) for i in range(n_rows)]
        for i, status in enumerate(project_summary["Status"]):
            point = series.data_points[i]
            point.graphicalProperties.solidFill = STATUS_COLORS_HEX[status]
            point.graphicalProperties.line.noFill = True

        worksheet.add_chart(chart, f"{get_column_letter(len(project_summary.columns) + 2)}2")

    return buffer.getvalue()


st.title("Tooling Data Cleaning & Budget Tool")

if st.text_input("Enter access code to continue", type="password") != st.secrets.get("site_access_code"):
    st.info("Enter the access code above to continue.")
    st.stop()

raw_sheets = None  # only (re)computed inside the change-gate below

if is_odbc_configured():
    st.caption(
        "Data is queried directly from Plex. Enter one Part No. filter per row "
        "below (e.g. 924) — add rows with the + at the bottom, or paste a whole "
        "column of them straight from Excel — then click 'Fetch from Plex'."
    )

    part_filters_df = st.data_editor(
        pd.DataFrame({"Part No. Filter": [""]}),
        key="part_filters_editor",
        num_rows="dynamic",
        hide_index=True,
        use_container_width=True,
    )
    fetch_clicked = st.button("Fetch from Plex")

    if fetch_clicked:
        part_filters = [
            str(v).strip() for v in part_filters_df["Part No. Filter"] if str(v).strip()
        ]
        if part_filters:
            with st.spinner(f"Querying Plex for {len(part_filters)} Part No. filter(s)..."):
                try:
                    st.session_state.plex_sheets = fetch_plex_sheets(part_filters)
                except Exception as e:
                    st.error(f"Couldn't query Plex: {e}")

    raw_sheets = st.session_state.get("plex_sheets", {})
    if not raw_sheets:
        st.info("Enter one or more Part No. filters above and click 'Fetch from Plex' to get started.")
        st.stop()
    sheet_key = list(raw_sheets.keys())
elif is_drive_configured():
    header_col, button_col = st.columns([4, 1])
    with header_col:
        st.caption("Files are fetched automatically from today's Plex export folder in Google Drive.")
    with button_col:
        if st.button("Refresh from Drive"):
            st.session_state.pop("drive_files", None)

    if "drive_files" not in st.session_state:
        with st.spinner("Fetching files from Google Drive..."):
            try:
                st.session_state.drive_files = fetch_drive_files()
            except Exception as e:
                st.error(f"Couldn't fetch files from Google Drive: {e}")
                st.session_state.drive_files = []

    uploaded_files = st.session_state.drive_files
    if not uploaded_files:
        st.info("No CSV/Excel files found in the configured Google Drive folder.")
        st.stop()
    sheet_key = [f.name for f in uploaded_files]
else:
    uploaded_files = st.file_uploader(
        "Upload CSV or Excel file(s)", type=["csv", "xlsx", "xls"], accept_multiple_files=True
    )
    if not uploaded_files:
        st.info("Upload a file to get started, or see PO_REGISTRY_SETUP.md to connect ODBC instead.")
        st.stop()
    sheet_key = [f.name for f in uploaded_files]

if st.session_state.get("loaded_sheet_key") != sheet_key:
    if raw_sheets is None:
        raw_sheets = load_all_sheets(uploaded_files)
    processed, missing_report = {}, {}

    for sheet_name, df in raw_sheets.items():
        missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
        if missing:
            missing_report[sheet_name] = missing
            continue
        df = df[df["Okay PN"].notna() & (df["Okay PN"].astype(str).str.strip() != "")]
        df = df[~df["Okay PN"].apply(is_junk_okay_pn)]
        df = df[~df["Okay PN"].map(normalize_pn).isin(EXCLUDED_OKAY_PNS)]
        processed[sheet_name] = consolidate_duplicates(df)

    po_registry = load_po_registry()
    st.session_state.registry_connected = is_registry_configured()

    st.session_state.loaded_sheet_key = sheet_key
    st.session_state.missing_report = missing_report
    st.session_state.editor_data = {
        name: df.assign(
            **{"Total PO $": df["Okay PN"].map(normalize_pn).map(po_registry).fillna(0.0)}
        )
        for name, df in processed.items()
    }
    st.session_state.pop("final_sheets", None)

for sheet_name, missing in st.session_state.missing_report.items():
    st.error(f"Sheet '{sheet_name}' is missing required columns {missing} — skipped.")

if not st.session_state.editor_data:
    st.stop()

st.success(f"Loaded {len(st.session_state.editor_data)} sheet(s) after cleaning duplicates.")
st.session_state.setdefault("registry_version", 0)

step2_header, step2_button = st.columns([4, 1])
with step2_header:
    st.header("Step 2: Total PO $ for each row")
if st.session_state.get("registry_connected"):
    with step2_button:
        if st.button("Refresh from registry"):
            fresh_registry = load_po_registry()
            for name, df in st.session_state.editor_data.items():
                st.session_state.editor_data[name] = df.assign(
                    **{"Total PO $": df["Okay PN"].map(normalize_pn).map(fresh_registry).fillna(df["Total PO $"])}
                )
            st.session_state.registry_version += 1
            st.rerun()
    st.caption(
        "'Total PO $' is locked here — it's pulled from your saved registry and can't be edited from "
        "this website. Update values in the 'PO Registry' tab of your Google Sheet, then click "
        "'Refresh from registry' to pull the latest values in without needing to re-upload."
    )
else:
    st.caption("'Total PO $' is locked and shown read-only — it can only be set via the Google Sheet registry.")
    st.warning(
        "PO $ registry isn't connected yet, so values aren't being remembered day to day. "
        "See PO_REGISTRY_SETUP.md to enable it."
    )

edited_data = {}
for sheet_name, df in st.session_state.editor_data.items():
    label = derive_sheet_label(df, sheet_name)
    with st.expander(label, expanded=len(st.session_state.editor_data) == 1):
        if label != sheet_name:
            st.caption(f"Source file: {sheet_name}")
        edited_data[sheet_name] = st.data_editor(
            df,
            key=f"editor_{sheet_name}_{st.session_state.registry_version}",
            disabled=True,
            use_container_width=True,
            num_rows="fixed",
        )

if st.button("Generate Final Results", type="primary"):
    final_sheets = {}
    for sheet_name, df in edited_data.items():
        df_final = df.copy()
        df_final["Budget Left"] = df_final["Total PO $"] - df_final["Total Cost"]
        final_sheets[sheet_name] = add_totals_row(df_final)
    st.session_state.final_sheets = final_sheets

    if st.session_state.get("registry_connected"):
        open_projects = build_open_projects(final_sheets)
        if log_open_projects_snapshot(open_projects):
            st.toast("Logged today's Project Balance snapshot to History.", icon="📈")

if "final_sheets" in st.session_state:
    tab_data, tab_report = st.tabs(["Data Cleaning", "Report"])

    with tab_data:
        st.header("Step 3: Final Results")
        for sheet_name, df_final in st.session_state.final_sheets.items():
            label = derive_sheet_label(df_final, sheet_name)
            st.subheader(label)
            if label != sheet_name:
                st.caption(f"Source file: {sheet_name}")
            st.dataframe(df_final, use_container_width=True)

        st.download_button(
            label="Download all sheets as Excel",
            data=build_excel(st.session_state.final_sheets),
            file_name="tooling_cleaned_all_sheets.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with tab_report:
        st.header("Step 4: Profit or Loss by Project")
        st.caption("One row per uploaded file, using its TOTAL row from Step 3. Profit or Loss = Total PO $ - Total Cost.")

        project_summary = build_project_summary(st.session_state.final_sheets).sort_values(
            "Profit or Loss ($)", ascending=False, ignore_index=True
        )

        total_open_projects = len(project_summary)
        projects_on_budget = int((project_summary["Status"] == "Profit").sum())
        pct_on_budget = (projects_on_budget / total_open_projects * 100) if total_open_projects else 0

        kpi1, kpi2, kpi3 = st.columns(3)
        kpi1.metric("Current % on Budget", f"{pct_on_budget:.0f}%")
        kpi2.metric("Projects On Budget", projects_on_budget)
        kpi3.metric("Total Open Projects", total_open_projects)

        st.dataframe(
            project_summary,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Total PO $": st.column_config.NumberColumn(format="$%.2f"),
                "Total Cost": st.column_config.NumberColumn(format="$%.2f"),
                "Profit or Loss ($)": st.column_config.NumberColumn(format="$%.2f"),
            },
        )

        st.download_button(
            label="Download project summary as Excel",
            data=build_project_summary_excel(project_summary),
            file_name="project_profit_or_loss_summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        project_order = project_summary["Project"].tolist()
        base_font = "system-ui, -apple-system, Segoe UI, sans-serif"

        bars = (
            alt.Chart(project_summary)
            .mark_bar(size=36, cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
            .encode(
                x=alt.X("Project:N", sort=project_order, title=None, axis=alt.Axis(labelAngle=0, labelColor="#52514e")),
                y=alt.Y(
                    "Profit or Loss ($):Q",
                    title="Profit or Loss ($)",
                    axis=alt.Axis(format="$,.0f", labelColor="#898781", gridColor="#e1e0d9", titleColor="#52514e"),
                ),
                color=alt.Color(
                    "Status:N",
                    scale=alt.Scale(domain=list(STATUS_COLORS.keys()), range=list(STATUS_COLORS.values())),
                    legend=alt.Legend(title=None, orient="top", symbolType="circle"),
                ),
                tooltip=[
                    "Project",
                    "File",
                    "Tooling Job No.",
                    alt.Tooltip("Total PO $:Q", format="$,.2f"),
                    alt.Tooltip("Total Cost:Q", format="$,.2f"),
                    alt.Tooltip("Profit or Loss ($):Q", format="$,.2f"),
                    "Status",
                ],
            )
        )
        labels = bars.mark_text(
            dy=alt.expr("datum['Profit or Loss ($)'] >= 0 ? -8 : 14"),
            color="#0b0b0b",
            fontSize=12,
            font=base_font,
        ).encode(text=alt.Text("Profit or Loss ($):Q", format="$,.0f"))
        zero_line = alt.Chart(pd.DataFrame({"y": [0]})).mark_rule(color="#c3c2b7", strokeWidth=1).encode(y="y:Q")

        chart = (
            (bars + zero_line + labels)
            .properties(
                height=380,
                title=alt.TitleParams(
                    "Profit or Loss by Project",
                    subtitle="Total PO $ minus Total Cost, per uploaded file",
                    fontSize=16,
                    subtitleFontSize=12,
                    subtitleColor="#898781",
                    anchor="start",
                    font=base_font,
                    subtitleFont=base_font,
                ),
            )
            .configure_view(strokeWidth=0)
            .configure_axis(labelFont=base_font, titleFont=base_font, grid=True, domain=False, tickSize=0)
            .configure_legend(labelFont=base_font, labelFontSize=12)
        )

        st.altair_chart(chart, use_container_width=True)

        st.header("Step 5: Open Projects")
        open_projects_code = st.text_input("Enter code to view Open Projects", type="password")
        if open_projects_code != st.secrets.get("open_projects_code", "4045"):
            st.info("Enter the access code above to view Open Projects.")
        else:
            st.caption(
                "Customer, Project, Part Number, Contingency/Management Reserve Used, Expected Project End Date "
                "and NOTES are maintained by hand in the 'Project Registry' tab of your Google Sheet. "
                "Project Balance is computed live: sum of Budget Left for that row's Part Number(s)."
            )

            if not st.session_state.get("registry_connected"):
                st.warning("PO $ registry isn't connected, so Open Projects can't be computed. See PO_REGISTRY_SETUP.md.")
            else:
                open_projects = build_open_projects(st.session_state.final_sheets)
                if open_projects.empty:
                    st.info(
                        "No rows yet in the 'Project Registry' tab. Add Customer / Project / Part Number rows "
                        "there and they'll show up here with Project Balance filled in automatically."
                    )
                else:
                    st.dataframe(
                        open_projects,
                        use_container_width=True,
                        hide_index=True,
                        column_config={"Project Balance": st.column_config.NumberColumn(format="$%.2f")},
                    )
